#!/usr/bin/env python3
"""disk_reconcile — reconcile a mirror drive against its catalog, both directions.

The existing three-drive verification system (10T / 8T / BLACK) hashes files
that exist on the drives and compares copy-to-copy. It is strong at "is this
copy faithful to that one?" and structurally blind to "is there something on
this drive nobody expected?" — an unlisted file is absent from the worklist
by definition, so a cross-drive walk never sees it. Three 28.8 GB temp files
once sat in __MASTERS__/_ARCHIVES_/ for six weeks, faithfully mirrored to
both clones, because nothing ever checked the drive against its catalog.

This tool checks a drive against ~/ob/kmr/SYS/SYS Catalog/Disk/Disk Master/
Master Contents.xlsx in both directions, plus two capacity questions:

  1. Expected-but-missing  — every catalog row: does its path exist on disk?
  2. Present-but-unexplained — every catalog-named container directory: does
     it hold anything the catalog doesn't list?
  3. Capacity ceiling       — can this drive's *capacity ceiling* (never its
     free space) still hold the source content?
  4. Refresh preview        — (only with --refresh-from) what would an
     rsync -aH --delete-before actually create / delete / transfer, and what
     are the 15 largest items it would spend headroom on?

Usage:
    disk_reconcile.py <DRIVE-LABEL> [--refresh-from <SOURCE-LABEL>] [--json]
                       [--root <path>] [--no-refresh-scan]

    --refresh-from <SOURCE-LABEL>  Also preview a refresh FROM this drive
                                    INTO <DRIVE-LABEL>, via rsync -n --stats.
    --json                         Emit one JSON object instead of text.
    --root <path>                  Override the mirror root registered for
                                    <DRIVE-LABEL> (testing against a fixture
                                    tree; production reads MIRROR_ROOTS).
    --no-refresh-scan              Skip the (slow — up to ~1hr on a full
                                    drive) rsync dry-run; direction 3 falls
                                    back to the catalog-derived estimate and
                                    direction 4 is reported as skipped.

Testing-support overrides (not part of the documented CLI above; needed to
exercise the tool against scratch fixtures without ever touching /Volumes):
    --catalog <path>          Override the Master Contents.xlsx path.
    --refresh-from-root <path>  Override the mirror root for --refresh-from's
                                 SOURCE-LABEL (mirrors --root, for the source
                                 side of a refresh preview).

Exit codes: 0 clean; 1 direction (1) missing, (2) unexplained, or
RELOCATED-CONTENTS-UNVERIFIED non-empty; 2 setup error (drive not mounted,
catalog unreadable, rsync failed).
"""

from __future__ import annotations

import argparse
import fnmatch
import json
import os
import plistlib
import re
import subprocess
import sys
import unicodedata
import zipfile
from typing import NoReturn
from dataclasses import dataclass, field
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write(
        "disk_reconcile requires openpyxl (pip install openpyxl). "
        "No fallback — install it rather than working around it.\n"
    )
    sys.exit(2)


# A clone's mirror root is NOT the volume root (BLACK holds the mirror in a
# subfolder). Unknown labels fall back to /Volumes/<LABEL> with a warning.
MIRROR_ROOTS = {
    "10T": "/Volumes/10T",
    "8T": "/Volumes/8T",
    "BLACK": "/Volumes/BLACK/Clone of 10T",
}

CATALOG_PATH = Path(
    "/Users/oblinger/ob/kmr/SYS/SYS Catalog/Disk/Disk Master/Master Contents.xlsx"
)
SHEET_NAME = "Master Contents"

# macOS per-volume system cruft — named explicitly so it gets a labeled
# suppression reason (not swallowed by the generic dotfile catch-all below).
MACOS_CRUFT_NAMES = {
    ".Spotlight-V100",
    ".fseventsd",
    ".DocumentRevisions-V100",
    ".Trashes",
    ".TemporaryItems",
    ".metadata_never_index",
    ".DS_Store",
}

# Ordered (matcher, reason) pairs for direction-2 (present-but-unexplained).
# First match wins. Every suppressed entry is still counted and reported —
# a guard that discards silently manufactures an invisible miss, which is
# exactly the failure mode this whole tool exists to catch elsewhere.
SUPPRESS_RULES = [
    (
        lambda n: fnmatch.fnmatchcase(n, "hashes-*.tsv") or fnmatch.fnmatchcase(n, "*.tsv"),
        "drive-local verification manifest (*.tsv) — per-drive hash ledger, not master content",
    ),
    (
        lambda n: n == "Master Contents.xlsx",
        "the catalog workbook — drive-local copy, not master content",
    ),
    (
        lambda n: n == "READ ME FIRST.txt",
        "drive-local orientation doc, not master content",
    ),
    (
        lambda n: fnmatch.fnmatchcase(n, "SYNC-LOG-*.txt"),
        "drive-local sync log, not master content",
    ),
    (
        lambda n: n in MACOS_CRUFT_NAMES,
        "macOS per-volume system cruft",
    ),
    (
        lambda n: n.startswith("."),
        "dotfile (macOS/editor artifact, not master content) — generic catch-all",
    ),
]

# The M3 disposition column does NOT carry a row-level "this row was relocated"
# flag, and the first version of this file assumed it did -- it matched the words
# "moved" / "reloc", which appear in ZERO of the 49 distinct values the real
# workbook actually holds. The predicate was green in tests only because the
# fixture invented disposition text the catalog never uses, so it could not have
# fired on real data: a probe that can never pass.
#
# What the column really encodes, per the workbook's own Legend sheet:
#   'M3 YYYY-MM-DD (R<n>/K<n>/M<n>)' -- PER-FILE counts within that row's subtree.
#   R = repaired in place, K = kept at canonical, M = moved out to `Broken {type}/`
#   inside `_ARCHIVES_/Master Reconciliation <date>/`. Only nonzero categories
#   appear, so '(R5/K12)' means zero moved.
#
# So M>0 does not mean the ROW moved -- it means n FILES under it did. A row is
# genuinely relocated only when EVERY file it contains was moved out, i.e. its
# M-count equals its Files count. (Both 2026-08-08 findings are single-file .zip
# rows stamped '(M1)': one file, one move, row gone.)
M3_COUNT_RE = re.compile(r"\b([RKME])(\d+)\b")

# The 10T Path cell may carry a curated trailing marker -- '[MOVED 2026-06-25]',
# '[NEVER LANDED]'. Two of the 83 rows do today, and both happen to be absent, so
# stripping changes no verdict right now; it is done anyway because the marker is
# an ANNOTATION about the row, not part of the path, and a future row annotated
# while still present would otherwise be reported missing purely because of its
# own label.
PATH_MARKER_RE = re.compile(r"\s*\[([^\]]+)\]\s*$")

# 'Never landed' is NOT a milder kind of missing and must not be filed with the
# relocated rows either -- it is the sharper finding of the two. The row asserts
# `C <date>` verification stamps for BOTH clones on content that was never on the
# master at all, which is the same defect ATT T147 records against the BEAST
# completeness sweep: a stamp reading as a measurement of something nobody
# checked. It gets its own section so it cannot be skimmed as bookkeeping.
NEVER_LANDED_RE = re.compile(r"never\s+landed", re.IGNORECASE)

# Where the moved bytes are supposed to have landed. Relocation is a CLAIM until
# this directory is confirmed to exist -- the design says "provided the bytes are
# findable at the recorded destination", and a stamp that vouches for an unchecked
# destination is the failure this whole feature exists to catch.
BROKEN_ARCHIVE_GLOB = "_ARCHIVES_/Master Reconciliation *"

# The destination directory existing is not evidence the bytes are inside it --
# a real 2026-06-25 finding was two rows reported "relocated, destination
# confirmed" on the strength of the directory alone, while their contents live
# inside a ZIP under it (`Broken Archives.zip`) that was never opened. Per the
# M3 legend, moved-out files land in `Broken <type>.zip` archives directly
# under the Master Reconciliation dir -- glob those and open each with
# zipfile.namelist() rather than trusting the directory listing.
BROKEN_ARCHIVE_ZIP_GLOB = "Broken *.zip"

# Anchored with a leading slash where the excluded name has real content
# nested *under* it on 10T (886 real files under nested .Spotlight-V100/
# dirs, 14 under nested .fseventsd/) — an unanchored --exclude matches the
# name at ANY depth and would silently drop those real files too. Only
# names with zero observed real-file content anywhere are left unanchored.
RSYNC_EXCLUDES = [
    "/.Spotlight-V100",
    "/.fseventsd",
    ".DocumentRevisions-V100",
    ".Trashes",
    ".TemporaryItems",
    ".metadata_never_index",
    "/hashes-*.tsv",
]


@dataclass
class CatalogRow:
    path: str
    bytes: int | None
    files: int | None
    stamp_8t: object
    stamp_black: object
    m3_disposition: object
    notes: object = None


def die(message: str, code: int) -> NoReturn:
    sys.stderr.write(f"disk_reconcile: {message}\n")
    sys.exit(code)


def parse_int_field(value):
    """Bytes/Files cells hold integers, but sometimes as strings, sometimes empty."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def gib(n):
    if n is None:
        return None
    return round(n / (1024 ** 3), 2)


def load_catalog(catalog_path: Path):
    if not catalog_path.exists():
        die(f"catalog not found: {catalog_path}", 2)
    wb = openpyxl.load_workbook(catalog_path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        die(f"sheet '{SHEET_NAME}' not found in {catalog_path}", 2)
    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        die(f"catalog sheet '{SHEET_NAME}' is empty: {catalog_path}", 2)
    col = {name: idx for idx, name in enumerate(header) if name is not None}
    required = ["10T Path", "Bytes", "Files", "8T", "BLACK", "M3 disposition"]
    missing_cols = [c for c in required if c not in col]
    if missing_cols:
        die(f"catalog missing expected column(s) {missing_cols} in {catalog_path}", 2)

    rows = []
    for raw in rows_iter:
        if raw is None:
            continue
        path = raw[col["10T Path"]]
        if path is None or str(path).strip() == "":
            continue
        rows.append(
            CatalogRow(
                path=str(path).strip(),
                bytes=parse_int_field(raw[col["Bytes"]]),
                files=parse_int_field(raw[col["Files"]]),
                stamp_8t=raw[col["8T"]],
                stamp_black=raw[col["BLACK"]],
                m3_disposition=raw[col["M3 disposition"]],
                notes=(raw[col["Notes"]] if "Notes" in col else None),
            )
        )
    return rows


def resolve_root(label: str, override: str | None) -> Path:
    if override:
        return Path(override)
    if label in MIRROR_ROOTS:
        return Path(MIRROR_ROOTS[label])
    sys.stderr.write(
        f"WARNING: no mirror root registered for drive label '{label}'; "
        f"falling back to /Volumes/{label}. Add it to MIRROR_ROOTS if this "
        "is a real drive.\n"
    )
    return Path(f"/Volumes/{label}")


def m3_counts(disposition):
    """Per-file R/K/M/E counts parsed out of an M3 disposition cell."""
    if not disposition:
        return {}
    return {k: int(v) for k, v in M3_COUNT_RE.findall(str(disposition))}


def strip_path_marker(path: str):
    """('__MASTERS__/x.zip [MOVED 2026-06-25]') -> ('__MASTERS__/x.zip', 'MOVED ...')"""
    m = PATH_MARKER_RE.search(path)
    if not m:
        return path, None
    return path[:m.start()].strip(), m.group(1).strip()


def never_landed(row, marker) -> bool:
    """Positive evidence the row was never on the master. Read from the curated
    marker or the Notes cell -- never inferred from absence, which is exactly the
    thing being explained."""
    if marker and NEVER_LANDED_RE.search(marker):
        return True
    return bool(row.notes and NEVER_LANDED_RE.search(str(row.notes)))


def _nfc(s: str) -> str:
    """macOS stores filenames decomposed (NFD) while the catalog is typed NFC,
    so 'Resumé' and 'Resumé' are the same file and compare unequal.
    Every name comparison against archive members goes through here."""
    return unicodedata.normalize("NFC", s)


def verify_archive_contents(dest: Path, real_path: str, expected_bytes=None,
                            expected_files=None):
    """Open every 'Broken *.zip' archive directly under dest and look for this
    row's file inside it. A directory existing is not evidence the bytes are
    inside it -- that is the entire difference between "the destination is
    there" and "the file is there".

    The match is GRADED, because the three strengths are not the same claim:

      path+size  the member's full catalogued path is present AND its declared
                 uncompressed size equals the row's Bytes. The strongest thing
                 obtainable without inflating 65 GB.
      path       the full catalogued path is present, but the row carries no
                 single-file byte count to compare against.
      basename   only the filename matched, at some other path. Reported as
                 confirmed but explicitly WEAK -- two files can share a name,
                 and a bare-name match is how an unrelated hit gets read as
                 proof (it happened during the F052 hunt: a substring search for
                 'Google Drive' matched unrelated Aeolus paths).

    Returns (state, detail):
      'confirmed'      - found; detail names which strength was achieved
      'size_mismatch'  - the path IS there but declares different bytes. This is
                         a finding, not a confirmation, and it outranks every
                         other failure in the report.
      'absent'         - readable zip(s) exist under dest, none list the file
      'no_archive'     - no 'Broken *.zip' archive exists under dest at all
      'unreadable'     - an archive exists but could not be opened (corrupt/IO)

    NOT verified here, and said out loud wherever this is reported: the declared
    sizes come from the zip central directory. No CRC is checked and nothing is
    inflated, so same-size corruption would pass. Calling that byte-verified
    would be the exact error this feature exists to catch.
    """
    archives = sorted(dest.glob(BROKEN_ARCHIVE_ZIP_GLOB))
    if not archives:
        return "no_archive", f"no '{BROKEN_ARCHIVE_ZIP_GLOB}' archive under {dest.name}/"
    want_path = _nfc(real_path.strip("/"))
    want_base = _nfc(Path(real_path).name)
    comparable = expected_bytes is not None and expected_files in (None, 1)
    checked = []
    weak_hit = None
    for archive in archives:
        try:
            with zipfile.ZipFile(archive) as zf:
                members = [(_nfc(i.filename), i.file_size) for i in zf.infolist()]
        except (zipfile.BadZipFile, OSError) as exc:
            return "unreadable", f"{archive.name} could not be opened: {exc}"
        checked.append(archive.name)
        for name, size in members:
            if name == want_path or name.endswith("/" + want_path):
                if not comparable:
                    return "confirmed", (
                        f"the full catalogued path is listed inside "
                        f"{archive.name} (path match; the row carries no "
                        f"single-file byte count to compare against). Declared "
                        f"sizes come from the zip central directory — no CRC "
                        f"was checked and nothing was inflated")
                if size == expected_bytes:
                    return "confirmed", (
                        f"the full catalogued path is listed inside "
                        f"{archive.name} declaring exactly the catalogued "
                        f"{expected_bytes:,} bytes (path + size match). "
                        f"Declared sizes come from the zip central directory — "
                        f"no CRC was checked and nothing was inflated, so "
                        f"same-size corruption would pass this check")
                return "size_mismatch", (
                    f"the catalogued path IS inside {archive.name}, but it "
                    f"declares {size:,} bytes where the catalog records "
                    f"{expected_bytes:,} — a difference of "
                    f"{size - expected_bytes:+,}. The bytes are not the bytes "
                    f"this row claims were preserved")
        if weak_hit is None:
            for name, size in members:
                if name == want_base or name.endswith("/" + want_base):
                    weak_hit = (archive.name, name, size)
                    break
    if weak_hit:
        aname, found_at, size = weak_hit
        return "confirmed", (
            f"WEAK MATCH — '{want_base}' appears inside {aname} at "
            f"'{found_at}', which is NOT the catalogued path '{want_path}'. A "
            f"filename can repeat; treat this as located-but-unproven rather "
            f"than as the row's bytes")
    return "absent", (
        f"'{want_path}' is not listed in {', '.join(checked)} — neither at its "
        f"catalogued path nor under its bare filename")


def relocation_verdict(row, root: Path, real_path: str):
    """Was this row's absence explained by the M3 move-out, and did the bytes
    actually land where the catalog says?

    Returns (kind, note) where kind is 'relocated' | 'contents_unverified' |
    'unconfirmed' | None. 'unconfirmed' is deliberately NOT treated as
    explained -- a row whose disposition suggests relocation but whose
    destination cannot be seen stays under missing, because the alternative
    is a tool that launders an unchecked claim into a clean report.
    'contents_unverified' is the sharper middle case: the destination
    DIRECTORY is there, but opening its archive did not confirm the bytes are
    actually inside it -- a distinction that exists only as a sentence in the
    report is not a fix, so it gets its own kind and its own non-zero exit."""
    moved = m3_counts(row.m3_disposition).get("M", 0)
    if not moved:
        return None, None
    if row.files is None:
        return "unconfirmed", (
            f"disposition moved {moved} file(s) out, but the row carries no "
            "Files count, so 'all of them' cannot be established")
    if moved < row.files:
        return "unconfirmed", (
            f"disposition moved {moved} of this row's {row.files} file(s) — a "
            "partial move does not explain the whole row being absent")
    dests = sorted(root.glob(BROKEN_ARCHIVE_GLOB))
    if not dests:
        return "unconfirmed", (
            f"all {moved} file(s) were moved out per the disposition, but no "
            f"'{BROKEN_ARCHIVE_GLOB}' exists on this drive to receive them")
    # Search EVERY Master Reconciliation dir, not just the first. The glob
    # carries a date wildcard precisely because a drive can accumulate more
    # than one pass, and a row archived by the second pass is confirmed just
    # as well as one archived by the first. Checking only dests[0] would turn
    # a second reconciliation date into a wave of false "contents unverified"
    # findings — each one exiting non-zero, which is the expensive direction
    # for a check whose whole value is that a finding means something.
    failures = []
    for dest in dests:
        state, detail = verify_archive_contents(
            dest, real_path, row.bytes, row.files)
        if state == "confirmed":
            return "relocated", (
                f"all {moved} file(s) moved to {dest.name}/ per the M3 pass; "
                f"contents confirmed — {detail}")
        failures.append((dest, state, detail))
    # No destination confirmed it. Report the loudest failure rather than the
    # first: a corrupt archive is a different problem from a merely absent
    # entry, and it must not be hidden behind an earlier dir that simply had
    # no archive in it.
    # size_mismatch outranks everything: "the path is there but the bytes are
    # different" is a louder finding than "it is not there at all", and it
    # must not be hidden behind a sibling directory that merely lacked an archive.
    rank = {"size_mismatch": 0, "unreadable": 1, "absent": 2, "no_archive": 3}
    dest, state, detail = min(failures, key=lambda f: rank[f[1]])
    return "contents_unverified", (
        f"all {moved} file(s) reportedly moved to {dest.name}/, but the "
        f"bytes could not be confirmed inside it — destination directory "
        f"present, contents unverified ({detail})")


def scan_missing(root: Path, catalog_rows: list[CatalogRow]):
    """Direction 1 — expected-but-missing (+ RELOCATED / contents-unverified split-out)."""
    missing = []
    relocated = []
    never = []
    unverified = []
    for row in catalog_rows:
        real_path, marker = strip_path_marker(row.path)
        if (root / real_path).exists():
            continue
        entry = {
            "path": row.path,
            "real_path": real_path,
            "marker": marker,
            "bytes": row.bytes,
            "gib": gib(row.bytes),
            "stamp_8t": row.stamp_8t,
            "stamp_black": row.stamp_black,
            "m3_disposition": row.m3_disposition,
        }
        entry["has_stale_stamp"] = bool(row.stamp_8t or row.stamp_black)
        # Never-landed is tested FIRST and wins outright. Both of the real rows
        # carry 'M3 ... (M1)', so an M-count test alone would file a file that
        # was never on the master as 'relocated, destination confirmed' -- the
        # tool would be manufacturing the exact false assurance it exists to find.
        if never_landed(row, marker):
            entry["why"] = (
                "recorded as never having landed on the master — so its 8T/BLACK "
                "stamps assert a verification of content this drive never held")
            never.append(entry)
            continue
        kind, note = relocation_verdict(row, root, real_path)
        entry["relocation_note"] = note
        if kind == "relocated":
            relocated.append(entry)
        elif kind == "contents_unverified":
            unverified.append(entry)
        else:
            missing.append(entry)
    for lst in (missing, relocated, never, unverified):
        lst.sort(key=lambda e: e["path"])
    return missing, relocated, never, unverified


def classify_suppression(name: str):
    for matcher, reason in SUPPRESS_RULES:
        if matcher(name):
            return reason
    return None


SIDECAR_SUFFIX = ".README.txt"


def sidecar_owner(name: str, expected_names) -> str | None:
    """A `<stem>.README.txt` beside a catalogued `<stem>.<ext>` is explained by
    that row, and returning its owner is what lets the report say so.

    This cannot live in SUPPRESS_RULES, which match a filename against a fixed
    pattern and nothing else.  A sidecar is not suppressible by name — a
    `.README.txt` with no catalogued sibling is a genuinely unexplained file and
    must stay in the report.  What explains it is the row it sits next to, which
    only the container's expected-name set knows.

    Written 2026-08-09 after the first clean run reported two sidecars written
    the previous night (F052's and F036's) as unexplained.  Suppressing them by
    a blanket `*.README.txt` rule would have muted the useful case along with
    the noise, which is the mute-button failure the exception list is supposed
    to avoid.
    """
    if not name.endswith(SIDECAR_SUFFIX):
        return None
    stem = name[: -len(SIDECAR_SUFFIX)]
    if not stem:
        return None
    for candidate in expected_names:
        if candidate == stem:
            return candidate
        base, dot, _ext = candidate.rpartition(".")
        if dot and base == stem:
            return candidate
    return None


def scan_unexplained(root: Path, catalog_rows: list[CatalogRow]):
    """Direction 2 — present-but-unexplained.

    Containers are the catalog-derived parent dirs of every row path, PLUS
    the drive root itself. The root is included even though no catalog row
    currently lives directly at depth-1: the suppression list (hashes-*.tsv,
    Master Contents.xlsx, READ ME FIRST.txt, SYNC-LOG-*.txt) names files
    that live at the drive root, and a purely catalog-container-derived scan
    would never look there, making that half of the exceptions list dead
    code. Scanning the root also catches a wholly novel top-level directory,
    which is in scope for "something nobody expected."

    A container can itself be the *parent* of another container (e.g. one
    catalog row is `__MASTERS__/__DAO_Readme.txt`, making `__MASTERS__` a
    scanned container — but `__MASTERS__` is also the parent of six other
    containers like `__MASTERS__/_ARCHIVES_`). Those intermediate
    directories are structurally required by deeper rows, not "unnamed by
    the catalog," so any entry that is itself one of the other scanned
    container paths is treated as expected regardless of which container
    it's being evaluated from.
    """
    containers: dict[str, set] = {}
    for row in catalog_rows:
        parts = row.path.split("/")
        container = "/".join(parts[:-1])
        containers.setdefault(container, set()).add(parts[-1])
    containers.setdefault("", set())
    all_container_paths = set(containers.keys())
    # ...and every ANCESTOR of a container, not only the containers themselves.
    # A container may sit several levels down with no row of its own at the
    # levels above it: with a single row `A/B/leaf`, the container is `A/B` and
    # `A` is nothing -- so `A` would be reported as an unexplained directory at
    # the drive root while being the only way to reach the catalogued content.
    # This does not bite on 10T today only because `__MASTERS__/__DAO_Readme.txt`
    # happens to make `__MASTERS__` a container in its own right; it would bite
    # on any drive whose catalog lacks such a row. Found by a fixture, not by 10T.
    for container in list(all_container_paths):
        parts = container.split("/") if container else []
        for depth in range(1, len(parts)):
            all_container_paths.add("/".join(parts[:depth]))

    unexplained = []
    suppressed: dict[str, list[str]] = {}
    for container in sorted(containers):
        disk_dir = root / container if container else root
        if not disk_dir.is_dir():
            continue
        expected_names = containers[container]
        for entry_name in sorted(os.listdir(disk_dir)):
            rel_path = f"{container}/{entry_name}" if container else entry_name
            if entry_name in expected_names or rel_path in all_container_paths:
                continue
            owner = sidecar_owner(entry_name, expected_names)
            reason = classify_suppression(entry_name)
            if owner:
                suppressed.setdefault(
                    f"sidecar README for a catalogued row — explained by `{owner}`", []
                ).append(rel_path)
            elif reason:
                suppressed.setdefault(reason, []).append(rel_path)
            else:
                unexplained.append(
                    {"container": container or "(drive root)", "name": entry_name, "path": rel_path}
                )
    unexplained.sort(key=lambda e: e["path"])
    return unexplained, suppressed


def catalog_bytes_sum(catalog_rows: list[CatalogRow]):
    return sum(r.bytes for r in catalog_rows if r.bytes)


def get_capacity_ceiling(root: Path):
    """Return (ceiling_bytes, method). Ceiling, never free space.

    Prefers `diskutil info -plist` (TotalSize) — the native macOS source of
    truth for a real /Volumes/<X> mountpoint. Falls back to os.statvfs
    (f_blocks * f_frsize) when diskutil can't resolve the path, which is
    always the case for a scratch/test fixture directory that isn't itself
    a distinct volume. statvfs's f_blocks is the filesystem's total block
    count — the ceiling — never affected by how much is currently free.
    """
    try:
        proc = subprocess.run(
            ["diskutil", "info", "-plist", str(root)],
            capture_output=True,
            timeout=30,
        )
        if proc.returncode == 0:
            plist = plistlib.loads(proc.stdout)
            total = plist.get("TotalSize")
            if isinstance(total, int) and total > 0:
                return total, "diskutil info -plist TotalSize"
    except (OSError, subprocess.SubprocessError, plistlib.InvalidFileException):
        pass
    st = os.statvfs(root)
    return (
        st.f_blocks * st.f_frsize,
        "os.statvfs f_blocks * f_frsize (diskutil unavailable for this path)",
    )


def build_rsync_cmd(src_root: Path, dst_root: Path):
    cmd = ["rsync", "-n", "-aH", "--delete-before", "--stats", "--human-readable", "--itemize-changes"]
    cmd += [f"--exclude={ex}" for ex in RSYNC_EXCLUDES]
    cmd += [f"{src_root}/", f"{dst_root}/"]
    return cmd


def parse_human_bytes(text: str):
    text = text.strip()
    m = re.match(r"^([\d,]+(?:\.\d+)?)\s*([KMGT]?)\s*$", text)
    if not m:
        return None
    num = float(m.group(1).replace(",", ""))
    mult = {"": 1, "K": 1_000, "M": 1_000_000, "G": 1_000_000_000, "T": 1_000_000_000_000}[m.group(2)]
    return int(round(num * mult))


def parse_refresh_output(output: str, src_root: Path):
    def find_int(pattern):
        m = re.search(pattern, output)
        return int(m.group(1).replace(",", "")) if m else None

    def find_bytes(pattern):
        m = re.search(pattern, output)
        return parse_human_bytes(m.group(1)) if m else None

    created = find_int(r"Number of created files:\s*([\d,]+)")
    deleted = find_int(r"Number of deleted files:\s*([\d,]+)")
    total_file_size = find_bytes(r"Total file size:\s*([\d,]+(?:\.\d+)?[KMGT]?)\s*bytes")
    total_transfer_size = find_bytes(r"Total transferred file size:\s*([\d,]+(?:\.\d+)?[KMGT]?)\s*bytes")

    top_files = []
    for line in output.splitlines():
        if len(line) <= 12:
            continue
        code = line[:11]
        if len(code) == 11 and code[1] == "f" and code[0] in "><ch.*":
            rel_path = line[12:]
            try:
                size = (src_root / rel_path).stat().st_size
            except OSError:
                continue
            top_files.append({"path": rel_path, "bytes": size, "gib": gib(size)})
    top_files.sort(key=lambda e: e["bytes"], reverse=True)

    return {
        "created": created,
        "deleted": deleted,
        "total_file_size": total_file_size,
        "total_transfer_size": total_transfer_size,
        "top_files": top_files[:15],
    }


def run_refresh_scan(src_root: Path, dst_root: Path):
    cmd = build_rsync_cmd(src_root, dst_root)
    sys.stderr.write(
        "Running rsync dry-run refresh scan — this can take up to an hour on a full "
        "drive (a prior full dry run over 8.3M files took 62 minutes).\n"
    )
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode not in (0, 24):  # 24: some source files vanished mid-scan (benign, dry-run)
        raise RuntimeError(f"rsync dry-run failed (exit {proc.returncode}): {proc.stderr.strip()[:500]}")
    return parse_refresh_output(proc.stdout, src_root)


def build_capacity_and_refresh(args, target_root: Path, catalog_rows: list[CatalogRow]):
    ceiling_bytes, ceiling_method = get_capacity_ceiling(target_root)
    refresh = None

    if args.refresh_from:
        source_root = resolve_root(args.refresh_from, args.refresh_from_root)
        if not source_root.exists():
            die(
                f"--refresh-from source root not found: {source_root} "
                f"(label '{args.refresh_from}'). Is it mounted?",
                2,
            )
        if args.no_refresh_scan:
            refresh = {"attempted": False, "skipped_reason": "--no-refresh-scan"}
            source_bytes = catalog_bytes_sum(catalog_rows)
            source_method = "catalog Bytes column sum (estimate; rsync scan skipped by --no-refresh-scan)"
        else:
            try:
                parsed = run_refresh_scan(source_root, target_root)
            except RuntimeError as exc:
                die(str(exc), 2)
            refresh = {
                "attempted": True,
                "source_root": str(source_root),
                "target_root": str(target_root),
                "files_created": parsed["created"],
                "files_deleted": parsed["deleted"],
                "total_transfer_bytes": parsed["total_transfer_size"],
                "total_transfer_gib": gib(parsed["total_transfer_size"]),
                "top15": parsed["top_files"],
            }
            source_bytes = parsed["total_file_size"]
            source_method = "rsync -n --stats 'Total file size' (measured source total, post-exclude)"
    else:
        source_bytes = catalog_bytes_sum(catalog_rows)
        source_method = "catalog Bytes column sum (estimate, not a live measurement)"

    fits = (source_bytes <= ceiling_bytes) if source_bytes is not None else None
    headroom = (ceiling_bytes - source_bytes) if source_bytes is not None else None
    capacity = {
        "target_root": str(target_root),
        "ceiling_bytes": ceiling_bytes,
        "ceiling_gib": gib(ceiling_bytes),
        "ceiling_method": ceiling_method,
        "source_bytes": source_bytes,
        "source_gib": gib(source_bytes),
        "source_method": source_method,
        "fits": fits,
        "headroom_bytes": headroom,
        "headroom_gib": gib(headroom),
    }
    return capacity, refresh


def render_text(result: dict) -> str:
    lines = []
    lines.append(f"=== disk reconcile: {result['drive']} ({result['root']}) ===")
    lines.append(f"catalog: {result['catalog_path']} ({result['catalog_rows']} rows)")
    lines.append("")

    lines.append(f"## (1) Expected-but-missing — {len(result['missing'])}")
    if result["missing"]:
        lines.append(
            "A verification stamp never expires. A row can read "
            '"verified present on both backups" while the path underneath it '
            "is empty — that false reassurance is exactly what this direction "
            "exists to catch."
        )
        for e in result["missing"]:
            gib_txt = f"{e['gib']} GiB" if e["gib"] is not None else "size unknown"
            lines.append(f"  - {e['path']}  ({gib_txt})")
            lines.append(f"      8T: {e['stamp_8t'] or '—'}   BLACK: {e['stamp_black'] or '—'}")
            lines.append(f"      M3 disposition: {e['m3_disposition'] or '—'}")
            if e.get("relocation_note"):
                lines.append(f"      relocation NOT confirmed: {e['relocation_note']}")
            if e["has_stale_stamp"]:
                lines.append(
                    "      ⚠ has a verification stamp despite being absent — "
                    "stamp does not mean the path still exists."
                )
    else:
        lines.append("  (none)")
    lines.append("")

    never = result["never_landed"]
    lines.append(f"## NEVER LANDED — {len(never)}")
    if never:
        lines.append(
            "Sharper than a missing row. These are catalogued as never having "
            "reached the master, yet they carry per-clone verification stamps — "
            "a stamp reading as a measurement of content no drive ever held."
        )
        for e in never:
            gib_txt = f"{e['gib']} GiB" if e["gib"] is not None else "size unknown"
            lines.append(f"  - {e['real_path']}  ({gib_txt})")
            lines.append(f"      8T: {e['stamp_8t'] or '—'}   BLACK: {e['stamp_black'] or '—'}")
            lines.append(f"      {e['why']}")
    else:
        lines.append("  (none)")
    lines.append("")

    lines.append(f"## RELOCATED (documented, not missing) — {len(result['relocated'])}")
    if result["relocated"]:
        for e in result["relocated"]:
            lines.append(f"  - {e['path']}  — {e['m3_disposition']}")
            if e.get("relocation_note"):
                lines.append(f"      {e['relocation_note']}")
    else:
        lines.append("  (none)")
    lines.append("")

    unverified = result["relocated_unverified"]
    lines.append(f"## RELOCATED — CONTENTS UNVERIFIED — {len(unverified)}")
    if unverified:
        lines.append(
            "Destination directory present but not opened -- the M3 pass's "
            "'Broken *.zip' archive either doesn't exist, couldn't be read, or "
            "doesn't list this row's file. A directory being there is not "
            "evidence the bytes are inside it; this is a finding, not bookkeeping."
        )
        for e in unverified:
            lines.append(f"  - {e['path']}  — {e['m3_disposition']}")
            if e.get("relocation_note"):
                lines.append(f"      {e['relocation_note']}")
    else:
        lines.append("  (none)")
    lines.append("")

    lines.append(f"## (2) Present-but-unexplained — {len(result['unexplained'])}")
    if result["unexplained"]:
        for e in result["unexplained"]:
            lines.append(f"  - {e['path']}  (in {e['container']})")
    else:
        lines.append("  (none)")
    lines.append("")

    lines.append("## Suppressed (exceptions — never silent)")
    if result["suppressed"]:
        for reason, paths in sorted(result["suppressed"].items()):
            lines.append(f"  - {reason}: {len(paths)}")
            for p in paths:
                lines.append(f"      {p}")
    else:
        lines.append("  (none suppressed)")
    lines.append("")

    cap = result["capacity"]
    lines.append("## (3) Capacity ceiling")
    lines.append(f"  target root:      {cap['target_root']}")
    lines.append(f"  ceiling:          {cap['ceiling_gib']} GiB  [{cap['ceiling_method']}]")
    lines.append(f"  source (bytes):   {cap['source_gib']} GiB  [{cap['source_method']}]")
    if cap["fits"] is None:
        lines.append("  fits:             unknown (no source size available)")
    else:
        verb = "fits" if cap["fits"] else "DOES NOT FIT"
        lines.append(f"  fits:             {verb}  (headroom: {cap['headroom_gib']} GiB)")
    lines.append("")

    lines.append("## (4) Refresh preview")
    refresh = result["refresh"]
    if refresh is None:
        lines.append("  (not requested — pass --refresh-from <SOURCE-LABEL>)")
    elif not refresh["attempted"]:
        lines.append(f"  skipped: {refresh['skipped_reason']}")
    else:
        lines.append(f"  {refresh['source_root']}  →  {refresh['target_root']}")
        lines.append(f"  files to create: {refresh['files_created']}")
        lines.append(f"  files to delete: {refresh['files_deleted']}")
        lines.append(
            f"  total transfer:  {refresh['total_transfer_gib']} GiB "
            f"({refresh['total_transfer_bytes']} bytes)"
        )
        lines.append("  top 15 largest items to be transferred:")
        for item in refresh["top15"]:
            lines.append(f"      {item['gib']} GiB  {item['path']}")

    return "\n".join(lines)


def render_json(result: dict) -> str:
    payload = {
        "missing": result["missing"],
        "relocated": result["relocated"],
        "relocated_unverified": result["relocated_unverified"],
        "never_landed": result["never_landed"],
        "unexplained": result["unexplained"],
        "capacity": result["capacity"],
        "refresh": result["refresh"],
        "suppressed": result["suppressed"],
    }
    return json.dumps(payload, indent=2, default=str)


def main():
    parser = argparse.ArgumentParser(
        description="Reconcile a mirror drive against its catalog, both directions, plus capacity."
    )
    parser.add_argument("drive", metavar="DRIVE-LABEL", help="10T, 8T, BLACK, or another registered label")
    parser.add_argument("--refresh-from", metavar="SOURCE-LABEL", default=None)
    parser.add_argument("--json", action="store_true")
    parser.add_argument("--root", default=None, help="Override the mirror root for DRIVE-LABEL")
    parser.add_argument(
        "--no-refresh-scan",
        action="store_true",
        help="Skip the rsync dry-run; direction 4 is skipped, direction 3 falls back to the catalog estimate",
    )
    parser.add_argument("--catalog", default=None, help="[testing] Override path to Master Contents.xlsx")
    parser.add_argument(
        "--refresh-from-root", default=None, help="[testing] Override the mirror root for --refresh-from"
    )
    args = parser.parse_args()

    catalog_path = Path(args.catalog) if args.catalog else CATALOG_PATH
    catalog_rows = load_catalog(catalog_path)

    target_root = resolve_root(args.drive, args.root)
    if not target_root.exists():
        die(f"drive root not found: {target_root} (expected mirror root for '{args.drive}'). Is it mounted?", 2)

    missing, relocated, never, relocated_unverified = scan_missing(target_root, catalog_rows)
    unexplained, suppressed = scan_unexplained(target_root, catalog_rows)
    capacity, refresh = build_capacity_and_refresh(args, target_root, catalog_rows)

    result = {
        "drive": args.drive,
        "root": str(target_root),
        "catalog_path": str(catalog_path),
        "catalog_rows": len(catalog_rows),
        "missing": missing,
        "relocated": relocated,
        "relocated_unverified": relocated_unverified,
        "never_landed": never,
        "unexplained": unexplained,
        "suppressed": suppressed,
        "capacity": capacity,
        "refresh": refresh,
    }

    if args.json:
        print(render_json(result))
    else:
        print(render_text(result))

    # A never-landed row is a finding, not bookkeeping: it means two clones carry
    # verification stamps for content the master never held. Excluding it from the
    # exit code would let the one class of defect this tool is best placed to see
    # exit 0 and be filed as clean. Same reasoning for relocated_unverified: a
    # destination directory existing is not confirmation the bytes are inside it,
    # and a distinction that only shows up as a sentence in the report -- never as
    # an exit code -- is not a fix, it's the defect being described.
    if missing or unexplained or never or relocated_unverified:
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()
